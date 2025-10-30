"""
Verify the Excel format matches legal standard
"""

from openpyxl import load_workbook

wb = load_workbook('wrongful_death_report_test.xlsx')

print("Worksheets:", wb.sheetnames)
print("\n" + "="*80)

# Read Summary sheet
ws = wb['Summary']

print("\nSUMMARY SHEET - First 25 rows:")
print("="*80)

for row in range(1, 26):
    row_data = []
    for col in range(1, 11):  # First 10 columns
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            row_data.append(f"Col{col}: {cell.value}")
    if row_data:
        print(f"Row {row}: {' | '.join(row_data)}")

print("\n" + "="*80)
print("\nYEARLY DATA TABLE (rows 16+):")
print("="*80)

# Show headers
header_row = 16
print(f"\nHeaders (Row {header_row}):")
for col in range(1, 11):
    cell = ws.cell(row=header_row, column=col)
    print(f"  Column {col}: {cell.value}")

# Show first 5 data rows
print(f"\nFirst 5 data rows:")
for row in range(17, 22):
    values = []
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        values.append(str(cell.value) if cell.value is not None else '')
    print(f"Row {row}: {values}")

print(f"\nTotal rows with data in Summary sheet: {ws.max_row}")
