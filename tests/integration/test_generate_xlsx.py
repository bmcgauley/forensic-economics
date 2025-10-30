"""
Integration test for XLSX generation workflow.

Tests the complete flow: intake -> agents -> aggregator -> XLSX.
"""

import pytest
import json
from pathlib import Path
from openpyxl import load_workbook

from src.models.intake import Intake
from src.agents.life_expectancy_agent import LifeExpectancyAgent
from src.agents.worklife_expectancy_agent import WorklifeExpectancyAgent
from src.agents.wage_growth_agent import WageGrowthAgent
from src.agents.discount_rate_agent import DiscountRateAgent
from src.agents.present_value_agent import PresentValueAgent
from src.aggregator import Aggregator
from src.xlsx.xlsx_generator import XLSXGenerator


@pytest.mark.integration
def test_complete_workflow(sample_intake, temp_dir):
    """Test complete workflow from intake to XLSX generation."""

    # 1. Validate intake
    intake = Intake(sample_intake)
    assert intake.victim_age == 35
    assert intake.salary == 85000.00

    # 2. Run agents
    agent_results = []

    # Life Expectancy Agent
    life_agent = LifeExpectancyAgent()
    life_result = life_agent.run(intake.to_dict())
    assert 'outputs' in life_result
    assert 'expected_remaining_years' in life_result['outputs']
    assert life_result['outputs']['expected_remaining_years'] > 0
    agent_results.append(life_result)

    # Worklife Expectancy Agent
    worklife_agent = WorklifeExpectancyAgent()
    worklife_result = worklife_agent.run(intake.to_dict())
    assert 'outputs' in worklife_result
    assert 'worklife_years' in worklife_result['outputs']
    assert worklife_result['outputs']['worklife_years'] > 0
    agent_results.append(worklife_result)

    # Wage Growth Agent
    wage_agent = WageGrowthAgent()
    wage_result = wage_agent.run(intake.to_dict())
    assert 'outputs' in wage_result
    assert 'annual_growth_rate' in wage_result['outputs']
    agent_results.append(wage_result)

    # Discount Rate Agent
    discount_agent = DiscountRateAgent()
    discount_result = discount_agent.run(intake.to_dict())
    assert 'outputs' in discount_result
    assert 'recommended_discount_rate' in discount_result['outputs']
    agent_results.append(discount_result)

    # Present Value Agent
    pv_input = {
        **intake.to_dict(),
        'worklife_years': worklife_result['outputs']['worklife_years'],
        'projected_wages': wage_result['outputs']['projected_wages_by_year'],
        'discount_curve': discount_result['outputs']['discount_curve']
    }
    pv_agent = PresentValueAgent()
    pv_result = pv_agent.run(pv_input)
    assert 'outputs' in pv_result
    assert 'total_present_value' in pv_result['outputs']
    assert pv_result['outputs']['total_present_value'] > 0
    agent_results.append(pv_result)

    # 3. Aggregate results
    aggregator = Aggregator()
    final_workbook = aggregator.aggregate(agent_results, intake.to_dict())

    assert 'summary' in final_workbook
    assert 'yearly' in final_workbook
    assert 'data_sources' in final_workbook
    assert 'methodology_notes' in final_workbook

    # Verify summary contains expected data
    summary = final_workbook['summary']
    assert 'victim_info' in summary
    assert summary['victim_info']['age'] == 35
    assert 'economic_summary' in summary
    assert summary['economic_summary']['total_present_value'] > 0

    # 4. Generate XLSX
    output_path = temp_dir / 'test_report.xlsx'
    xlsx_generator = XLSXGenerator()
    result_path = xlsx_generator.generate(final_workbook, str(output_path))

    # Verify file was created
    assert Path(result_path).exists()

    # 5. Validate XLSX structure
    wb = load_workbook(result_path)

    # Check required worksheets exist
    expected_sheets = ['Summary', 'Yearly Detail', 'Data Sources', 'Methodology']
    for sheet_name in expected_sheets:
        assert sheet_name in wb.sheetnames, f"Missing worksheet: {sheet_name}"

    # Verify Summary sheet has data
    summary_ws = wb['Summary']
    assert summary_ws['A1'].value == "WRONGFUL DEATH ECONOMIC LOSS SUMMARY"

    # Verify Yearly Detail sheet has headers
    yearly_ws = wb['Yearly Detail']
    assert yearly_ws['A1'].value == "Year"

    # Close workbook
    wb.close()


@pytest.mark.integration
def test_sample_intake_file(temp_dir):
    """Test using the sample intake JSON file."""

    # Load sample intake file
    sample_file = Path('specs/1-wrongful-death-econ/samples/sample_intake.json')

    # Check if file exists, if not create it
    if not sample_file.exists():
        sample_data = {
            "id": "sample-001",
            "victim_age": 42,
            "victim_sex": "M",
            "occupation": "Construction Manager",
            "education": "bachelors",
            "salary": 95000.00,
            "salary_type": "current",
            "location": "CA",
            "dependents": 3,
            "benefits": {
                "retirement_contribution": 9500.00,
                "health_benefits": 12000.00
            },
            "metadata": {
                "submission_timestamp": "2025-10-29T12:00:00Z",
                "notes": "Sample intake for testing"
            }
        }
        sample_file.parent.mkdir(parents=True, exist_ok=True)
        with open(sample_file, 'w') as f:
            json.dump(sample_data, f, indent=2)

    # Load and validate
    with open(sample_file, 'r') as f:
        data = json.load(f)

    intake = Intake(data)
    assert intake.victim_age > 0
    assert intake.salary > 0


@pytest.mark.integration
def test_provenance_tracking(sample_intake):
    """Test that provenance is tracked throughout the workflow."""

    intake = Intake(sample_intake)

    # Run life expectancy agent
    life_agent = LifeExpectancyAgent()
    life_result = life_agent.run(intake.to_dict())

    # Verify provenance log exists
    assert 'provenance_log' in life_result
    assert len(life_result['provenance_log']) > 0

    # Verify provenance entries have required fields
    for entry in life_result['provenance_log']:
        assert 'step' in entry
        assert 'description' in entry
        assert 'value' in entry
        assert 'timestamp' in entry or 'source_date' in entry
