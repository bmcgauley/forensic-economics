#!/usr/bin/env python
"""
MVP Verification Script

Tests the complete workflow to ensure the MVP is working correctly.
"""

import sys
import os
import json
from pathlib import Path

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))

from src.models.intake import Intake
from src.agents.life_expectancy_agent import LifeExpectancyAgent
from src.agents.worklife_expectancy_agent import WorklifeExpectancyAgent
from src.agents.wage_growth_agent import WageGrowthAgent
from src.agents.discount_rate_agent import DiscountRateAgent
from src.agents.present_value_agent import PresentValueAgent
from src.aggregator import Aggregator
from src.xlsx.xlsx_generator import XLSXGenerator


def print_header(text):
    """Print a formatted header."""
    print("\n" + "=" * 70)
    print(f"  {text}")
    print("=" * 70)


def print_success(text):
    """Print success message."""
    print(f"[OK] {text}")


def print_error(text):
    """Print error message."""
    print(f"[ERROR] {text}")


def verify_mvp():
    """Verify the MVP is working end-to-end."""

    print_header("Forensic Economics - MVP Verification")

    # 1. Load sample intake
    print("\n1. Loading sample intake...")
    sample_file = Path('specs/1-wrongful-death-econ/samples/sample_intake.json')

    if not sample_file.exists():
        print_error(f"Sample file not found: {sample_file}")
        return False

    with open(sample_file, 'r') as f:
        intake_data = json.load(f)

    print_success(f"Loaded sample intake for {intake_data.get('occupation')}")

    # 2. Validate intake
    print("\n2. Validating intake data...")
    try:
        intake = Intake(intake_data)
        print_success(f"Intake validated: Age {intake.victim_age}, Salary ${intake.salary:,.2f}")
    except Exception as e:
        print_error(f"Intake validation failed: {e}")
        return False

    # 3. Run agents
    print("\n3. Running calculation agents...")
    agent_results = []

    try:
        # Life Expectancy
        print("   - Life Expectancy Agent...", end=" ")
        life_agent = LifeExpectancyAgent()
        life_result = life_agent.run(intake.to_dict())
        agent_results.append(life_result)
        print(f"[OK] {life_result['outputs']['expected_remaining_years']:.1f} years remaining")

        # Worklife Expectancy
        print("   - Worklife Expectancy Agent...", end=" ")
        worklife_agent = WorklifeExpectancyAgent()
        worklife_result = worklife_agent.run(intake.to_dict())
        agent_results.append(worklife_result)
        print(f"[OK] {worklife_result['outputs']['worklife_years']:.1f} worklife years")

        # Wage Growth
        print("   - Wage Growth Agent...", end=" ")
        wage_agent = WageGrowthAgent()
        wage_result = wage_agent.run(intake.to_dict())
        agent_results.append(wage_result)
        print(f"[OK] {wage_result['outputs']['annual_growth_rate']:.2%} growth rate")

        # Discount Rate
        print("   - Discount Rate Agent...", end=" ")
        discount_agent = DiscountRateAgent()
        discount_result = discount_agent.run(intake.to_dict())
        agent_results.append(discount_result)
        print(f"[OK] {discount_result['outputs']['recommended_discount_rate']:.2%} discount rate")

        # Present Value
        print("   - Present Value Agent...", end=" ")
        pv_input = {
            **intake.to_dict(),
            'worklife_years': worklife_result['outputs']['worklife_years'],
            'projected_wages': wage_result['outputs']['projected_wages_by_year'],
            'discount_curve': discount_result['outputs']['discount_curve']
        }
        pv_agent = PresentValueAgent()
        pv_result = pv_agent.run(pv_input)
        agent_results.append(pv_result)
        print(f"[OK] ${pv_result['outputs']['total_present_value']:,.2f} total PV")

    except Exception as e:
        print_error(f"Agent execution failed: {e}")
        import traceback
        traceback.print_exc()
        return False

    # 4. Aggregate results
    print("\n4. Aggregating results...")
    try:
        aggregator = Aggregator()
        final_workbook = aggregator.aggregate(agent_results, intake.to_dict())
        print_success("Results aggregated successfully")
        print(f"   - {len(final_workbook['yearly'])} yearly entries")
        print(f"   - {len(final_workbook['data_sources'])} data sources tracked")
    except Exception as e:
        print_error(f"Aggregation failed: {e}")
        import traceback
        traceback.print_exc()
        return False

    # 5. Generate XLSX
    print("\n5. Generating Excel workbook...")
    try:
        output_dir = Path('temp')
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / 'test_report.xlsx'

        xlsx_generator = XLSXGenerator()
        result_path = xlsx_generator.generate(final_workbook, str(output_path))

        print_success(f"Excel workbook generated: {result_path}")
        print(f"   - File size: {Path(result_path).stat().st_size:,} bytes")

        # Verify workbook can be read
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        print(f"   - Worksheets: {', '.join(wb.sheetnames)}")
        wb.close()

    except Exception as e:
        print_error(f"XLSX generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False

    # Summary
    print_header("Verification Complete")
    print("\n[OK] All components working correctly!")
    print(f"\nGenerated report: {output_path}")
    print("\nYou can now:")
    print("  1. Run 'python run.py' to start the web server")
    print("  2. Open http://localhost:5000 in your browser")
    print("  3. Submit intake data and generate reports")
    print("\nMVP is ready for use!")

    return True


if __name__ == '__main__':
    try:
        success = verify_mvp()
        sys.exit(0 if success else 1)
    except KeyboardInterrupt:
        print("\n\nVerification cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n\nUnexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
