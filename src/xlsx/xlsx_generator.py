"""
XLSX Generator Module

Purpose: Consume FinalWorkbook JSON and generate legally-compatible Excel workbook.
Produces workbook with Summary, Yearly Detail, Data Sources, and Methodology sheets.

Single-file module (target <=260 lines)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from datetime import datetime
import os


class XLSXGenerator:
    """Generate Excel workbooks from aggregated calculation results."""

    def __init__(self):
        """Initialize generator with style definitions."""
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self, final_workbook: Dict[str, Any], output_path: str) -> str:
        """
        Generate Excel workbook from final workbook data.

        Args:
            final_workbook: FinalWorkbook dictionary from aggregator
            output_path: Path where to save the workbook

        Returns:
            Path to generated workbook
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create worksheets
        self._create_summary_sheet(wb, final_workbook)
        self._create_yearly_detail_sheet(wb, final_workbook)
        self._create_data_sources_sheet(wb, final_workbook)
        self._create_methodology_sheet(wb, final_workbook)

        # Save workbook
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create Summary worksheet with legal standard format."""
        ws = wb.create_sheet("Summary", 0)
        summary = data.get('summary', {})
        yearly = data.get('yearly', [])
        victim_info = summary.get('victim_info', {})
        econ = summary.get('economic_summary', {})

        # Get cumulative present value from last row of yearly data
        cumulative_pv = yearly[-1].get('cumulative_present_value', 0) if yearly else 0

        row = 1

        # Title
        ws[f'A{row}'] = "WRONGFUL DEATH ECONOMIC LOSS SUMMARY"
        ws[f'A{row}'].font = self.title_font
        row += 2

        # Header section - victim information and key values
        ws[f'A{row}'] = "Name:"
        ws[f'B{row}'] = victim_info.get('full_name', '[CONFIDENTIAL]')
        row += 1

        ws[f'A{row}'] = "Item:"
        ws[f'B{row}'] = "Earnings Loss (More Conservative)"
        ws[f'B{row}'].font = Font(color="C65D57")  # Reddish color like in image
        row += 2

        # Key values in right-aligned format
        ws[f'F{row}'] = econ.get('current_salary', 0)
        ws[f'F{row}'].number_format = '$#,##0.00'
        ws[f'F{row}'].font = Font(color="C65D57")
        ws[f'G{row}'] = "<-- Base Value"
        row += 1

        ws[f'F{row}'] = econ.get('discount_rate', 0)
        ws[f'F{row}'].number_format = '0.00%'
        ws[f'F{row}'].font = Font(color="C65D57")
        ws[f'G{row}'] = "<-- Discount rate"
        row += 1

        ws[f'F{row}'] = econ.get('wage_growth_rate', 0)
        ws[f'F{row}'].number_format = '0.00%'
        ws[f'F{row}'].font = Font(color="C65D57")
        ws[f'G{row}'] = "<-- Annual growth rate"
        row += 2

        ws[f'F{row}'] = cumulative_pv
        ws[f'F{row}'].number_format = '$#,##0'
        ws[f'F{row}'].font = Font(color="C65D57")
        ws[f'G{row}'] = "<-- Cumulative Present Value"
        row += 2

        # Present Value Date
        ws[f'E{row}'] = "Present Value Date:"
        ws[f'E{row}'].font = self.header_font
        row += 1

        # Get present date from version metadata or use current date
        from datetime import datetime
        present_date_str = data.get('version_metadata', {}).get('created_at', datetime.utcnow().isoformat())
        try:
            present_date = datetime.fromisoformat(present_date_str.replace('Z', '+00:00'))
        except:
            present_date = datetime.utcnow()

        ws[f'E{row}'] = "Month -->"
        ws[f'F{row}'] = present_date.strftime('%b')
        row += 1
        ws[f'E{row}'] = "Day -->"
        ws[f'F{row}'] = present_date.day
        row += 1
        ws[f'E{row}'] = "Year -->"
        ws[f'F{row}'] = present_date.year
        row += 3

        # Table headers - legal standard format
        headers = [
            'Age',
            'Start\nDate',
            'Year\nNumber',
            'Portion\nof Year',
            'Full Year\nValue',
            'Actual\nValue',
            'Cumulative\nValue',
            'Discount\nFactor',
            'Present\nValue',
            'Cumulative\nPresent Value'
        ]

        header_row = row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        row += 1

        # Data rows - populate from yearly cashflows
        for year_data in yearly:
            age = year_data.get('age', 0)

            ws.cell(row=row, column=1, value=age).number_format = '0.0'
            ws.cell(row=row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
            ws.cell(row=row, column=2, value=year_data.get('start_date', '')).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=year_data.get('year_number', 0)).number_format = '0.0'
            ws.cell(row=row, column=4, value=year_data.get('portion_of_year', 0)).number_format = '0.00'
            ws.cell(row=row, column=5, value=year_data.get('full_year_value', 0)).number_format = '$#,##0'
            ws.cell(row=row, column=6, value=year_data.get('actual_value', 0)).number_format = '$#,##0'
            ws.cell(row=row, column=7, value=year_data.get('cumulative_value', 0)).number_format = '$#,##0'
            ws.cell(row=row, column=8, value=year_data.get('discount_factor', 0)).number_format = '0.00000'
            ws.cell(row=row, column=9, value=year_data.get('present_value', 0)).number_format = '$#,##0'
            ws.cell(row=row, column=10, value=year_data.get('cumulative_present_value', 0)).number_format = '$#,##0'

            # Add borders to all cells in this row
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = self.border

            row += 1

        # Adjust column widths for readability
        ws.column_dimensions['A'].width = 8   # Age
        ws.column_dimensions['B'].width = 10  # Start Date
        ws.column_dimensions['C'].width = 8   # Year Number
        ws.column_dimensions['D'].width = 10  # Portion of Year
        ws.column_dimensions['E'].width = 12  # Full Year Value
        ws.column_dimensions['F'].width = 12  # Actual Value
        ws.column_dimensions['G'].width = 15  # Cumulative Value
        ws.column_dimensions['H'].width = 12  # Discount Factor
        ws.column_dimensions['I'].width = 12  # Present Value
        ws.column_dimensions['J'].width = 18  # Cumulative Present Value

    def _create_yearly_detail_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create Yearly Detail worksheet with year-by-year calculations."""
        ws = wb.create_sheet("Yearly Detail")
        yearly = data.get('yearly', [])

        # Headers
        headers = ['Year', 'Age', 'Base Wage', 'Total Compensation', 'Discount Rate', 'PV Factor', 'Present Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, year_data in enumerate(yearly, 2):
            ws.cell(row=row_idx, column=1, value=year_data.get('year', ''))
            ws.cell(row=row_idx, column=2, value=year_data.get('age', ''))
            ws.cell(row=row_idx, column=3, value=year_data.get('base_wage', 0)).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=4, value=year_data.get('total_compensation', 0)).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=5, value=year_data.get('discount_rate', 0)).number_format = '0.00%'
            ws.cell(row=row_idx, column=6, value=year_data.get('pv_factor', 0)).number_format = '0.000000'
            ws.cell(row=row_idx, column=7, value=year_data.get('present_value', 0)).number_format = '$#,##0.00'

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_data_sources_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create Data Sources worksheet."""
        ws = wb.create_sheet("Data Sources")
        data_sources = data.get('data_sources', [])

        # Title
        ws['A1'] = "DATA SOURCES & PROVENANCE"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        # Headers
        headers = ['Agent', 'Source Name', 'URL', 'Usage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Data rows
        for row_idx, source in enumerate(data_sources, 4):
            ws.cell(row=row_idx, column=1, value=source.get('agent', ''))
            ws.cell(row=row_idx, column=2, value=source.get('source_name', ''))
            ws.cell(row=row_idx, column=3, value=source.get('source_url', ''))
            ws.cell(row=row_idx, column=4, value=source.get('usage', ''))

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25

    def _create_methodology_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create Methodology worksheet."""
        ws = wb.create_sheet("Methodology")

        methodology_notes = data.get('methodology_notes', '')

        # Title
        ws['A1'] = "CALCULATION METHODOLOGY"
        ws['A1'].font = self.title_font

        # Add methodology text
        ws['A3'] = methodology_notes
        ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')

        # Adjust column width
        ws.column_dimensions['A'].width = 100
        ws.row_dimensions[3].height = 400
